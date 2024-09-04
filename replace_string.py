import os
import pandas as pd
import zipfile
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

def replace_string_in_excel_files(directory, search_string, replace_string):
    for filename in os.listdir(directory):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(directory, filename)
            
            try:
                print(f"Processing {filename}")
                
                # Attempt to read the Excel file
                # xls = pd.ExcelFile(file_path)
                workbook = load_workbook(file_path)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # Iterate through each row and cell
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str):
                                # Replace the string in the cell value
                                cell.value = cell.value.replace(search_string, replace_string)
                
                # Save the modified workbook
                workbook.save(file_path)
                # # Ensure the file has visible sheets
                # if not xls.sheet_names:
                #     print(f"Skipping {filename}: No visible sheets found.")
                #     continue

                # # Process each sheet in the Excel file
                # with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                #     for sheet_name in workbook.sheet_names:
                #         df = pd.read_excel(workbook, sheet_name=sheet_name)
                #         df = df.replace(search_string, replace_string, regex=False)
                #         df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                print(f"Processed {filename}")
            
            except (zipfile.BadZipFile, InvalidFileException) as e:
                print(f"Skipping {filename}: Corrupted or invalid file.")
            except IndexError as e:
                print(f"Skipping {filename}: {e}")
            except Exception as e:
                print(f"Error processing {filename}: {e}")

# Define the directory containing the Excel files
directory = 'C:/Users/BRSBAWAM/Desktop/mapweb'  # Change this to your directory path

# Define the string to search and replace
search_string = '$0003($FLD32)'
replace_string = '$0003($FLD15)'

# $0001($FLD44)

# Run the function
replace_string_in_excel_files(directory, search_string, replace_string)
